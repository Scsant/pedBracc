import os
from openpyxl import Workbook, load_workbook
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

# Caminho do arquivo Excel
EXCEL_FILE_PATH = "/tmp/vale_pedagio.xlsx" 

def salvar_dados_excel(dados):
    # Verificar se o arquivo existe
    if os.path.exists(EXCEL_FILE_PATH):
        workbook = load_workbook(EXCEL_FILE_PATH)
        sheet = workbook.active
    else:
        # Criar um novo arquivo Excel com cabeçalhos
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "Data da Emissão", "Placa do Caminhão", "Fazenda Destino",
            "Número Recibo Ida", "Número Recibo Volta",
            "Custo Pedágio Ida", "Custo Pedágio Volta",
            "Custo Pedágio Total", "Coordenadas X", "Coordenadas Y"
        ])

    # Adicionar os dados na próxima linha
    sheet.append([
        dados.get("data_emissao"),
        dados.get("placa_caminhao"),
        dados.get("fazenda_destino"),
        dados.get("numero_recibo_ida"),
        dados.get("numero_recibo_volta"),
        dados.get("custo_pedagio_ida"),
        dados.get("custo_pedagio_volta"),
        dados.get("custo_pedagio_total"),
        dados.get("coordenadas_x", 0.0),  # Valor padrão
        dados.get("coordenadas_y", 0.0),  # Valor padrão
    ])

    # Salvar o arquivo
    workbook.save(EXCEL_FILE_PATH)

class RegistroValePedagioAPIView(APIView):
    def post(self, request):
        # Dados enviados no corpo da requisição
        dados = request.data

        # Salvar no Excel
        try:
            salvar_dados_excel(dados)
            return Response({"mensagem": "Dados salvos com sucesso!"}, status=status.HTTP_201_CREATED)
        except Exception as e:
            return Response({"erro": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


from django.http import FileResponse

class DownloadExcelAPIView(APIView):
    def get(self, request):
        if os.path.exists(EXCEL_FILE_PATH):
            return FileResponse(open(EXCEL_FILE_PATH, 'rb'), as_attachment=True, filename="vale_pedagio.xlsx")
        return Response({"erro": "Arquivo não encontrado"}, status=404)

